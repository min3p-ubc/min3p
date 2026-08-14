########################################################################
#
# Authors: 
#   Danyang Su
#
# Description:
#   CDR related result post-processing using python script
#
# (c) Copyright:
#   Terradot Soil, Inc.
#
# Dependency:
#   simpleeval, openpyxl
#
########################################################################

# Description of CDR input file (example): 
# !**********************************************************************
# !CDR Template File
# !
# !Author:
# !  Danyang Su
# !
# !Description:
# !  This script processes MIN3P simulation results to extract and calculate
# !  CDR (Carbon Dioxide Removal) related data using Python, then exports 
# !  the results to an Excel file.
# !
# !  The input file follows a structured format organized into groups:
# !  1. Each group begins with the line group: [name] and ends with end group
# !  2. Lines starting with '!' are treated as comments.
# !
# !  Within each group:
# !  The first line specifies the number of output columns,
# !  Subsequent lines define each variable with:
# !  1. A description
# !  2. The output variable name (for Excel)
# !  3. The MIN3P result file extension
# !  4. The corresponding variable name in MIN3P result file
# !
# !  The script supports basic mathematical operations for unit conversions 
# !  and data adjustments. For example:
# !  1. To convert CO₂ values from mol/m² to tn/Ha, multiply by a coefficient of 0.44
# !  2. Some results require sign inversion for correct interpretation.
# !**********************************************************************
# 
# 'group: co2'             ;keyword for CDR output section
# 15                       ;number of output columns
# 'v1=v1*(-0.44)'          'Autotrophic soil respiration [tn CO2/Ha]'    'o.arup'    'co3-2 accumulative solute uptake [mol]'
# 'v2=v2*0.44'             'Heterotrophic soil respiration [tn CO2/Ha]'  '10.mmac'   'accumulative source/sink from ch2o-o2-td [mol]'
# 'v3=v3*0.44'             'UAN hydrolisis [tn CO2/Ha]'                  '10.mmac'   'accumulative source/sink from uan-fertilizer-n [mol]'
# 'v4=v4*0.44'             'Urea hydrolisis (PO4) [tn CO2/Ha]'           '10.mmac'   'accumulative source/sink from urea-phosphate [mol]'
# 'v5=v5*0.44'             'CO2 in calcite (in) [tn CO2/Ha]'             '10.mmac'   'accumulative source/sink from calcite [mol]'
# 'v6=v6*0.44'             'CO2 inflow [tn CO2/Ha]'                      '10.mac'    'accumulative mass influx [mol]'
# 'v7=v7*0.44'             'CO2 outflux [tn CO2/Ha]'                     '10.mac'    'accumulative mass outflux (gas phase) [mol]'
# 'v8=v8*0.44'             'CO2 export [tn CO2/Ha]'                      '10.mac'    'accumulative mass outflux [mol]'
# 'v9=v9*0.44'             'CO2 change storage (aqueous) [tn CO2/Ha]'    '10.mac'    'accumulative change in storage [mol]'
# 'v10=v10*0.44'           'CO2 change storage (gas) [tn CO2/Ha]'        '10.mac'    'accumulative change in storage (gas phase) [mol]'
# 'v11=v1+v2+v3+v4+v5+v6'  'CO2 source [tn CO2/Ha]'                      ''          ''
# 'v12=v7+v8'              'CO2 sink [tn CO2/Ha]'                        ''          ''
# 'v13=v9+v10'             'CO2 change storage [tn CO2/Ha]'              ''          ''
# 'v14=v11-v12-v13'        'CO2 error balance [tn CO2/Ha]'               ''          ''
# 'v15=v14/v11*100.0'      'CO2 error balance [%]'                       ''          ''
# 'end group'
# 
# 'group: ca'              ;keyword for CDR output section
# 9                        ;number of output columns
# 'v1=v1'                  'Weathered [mol/m2]'                          '1.mmac'    'accumulative source from all minerals [mol]'
# 'v2=max(v2,0.0)'         'Calcite diss. [mol/m2]'                      '1.mmac'    'accumulative source/sink from calcite [mol]'    ;> 0
# 'v3=max(-v3,0.0)'        'Calcite prec. [mol/m2]'                      '1.mmac'    'accumulative source/sink from calcite [mol]'    ;< 0
# 'v4=v4'                  'Turnover [mol/m2]'                           '1.mmac'    'accumulative source/sink from turnover-26-12-2024 [mol]'
# 'v5=v5*(-1.0)'           'Exchanged [mol/m2]'                          '1.mmac'    'accumulative source/sink from ion-exchange-ca [mol]'
# 'v6=v6'                  'Root uptake [mol/m2]'                        'o.rup'     'ca+2 accumulative solute uptake [mol]'
# 'v7=v7'                  'Influx [mol/m2]'                             '1.mac'     'accumulative mass influx [mol]'
# 'v8=v8'                  'Exported [mol/m2]'                           '1.mac'     'accumulative mass outflux [mol]'
# 'v9=v9'                  'Change storage (aq) [mol/m2]'                '1.mac'     'accumulative change in storage [mol]'
# 'end group'
# 
# 'group: mg'              ;keyword for CDR output section
# 7                        ;number of output columns
# 'v1=v1'                  'Weathered [mol/m2]'                          '2.mmac'    'accumulative source from all minerals [mol]'
# 'v2=v2'                  'Turnover [mol/m2]'                           '2.mmac'    'accumulative source/sink from turnover-26-12-2024 [mol]'
# 'v3=v3'                  'Exchanged [mol/m2]'                          '2.mmac'    'accumulative source/sink from ion-exchange-ca [mol]'
# 'v4=v4'                  'Root uptake [mol/m2]'                        'o.rup'     'mg+2 accumulative solute uptake [mol]'
# 'v5=v5'                  'Influx [mol/m2]'                             '2.mac'     'accumulative mass influx [mol]'
# 'v6=v6'                  'Exported [mol/m2]'                           '2.mac'     'accumulative mass outflux [mol]'
# 'v7=v7'                  'Change storage (aq) [mol/m2]'                '2.mac'     'accumulative change in storage [mol]'
# 'end group'
########################################################################

import sys
import math
import ast
import operator
import re
import numpy as np
from openpyxl import Workbook 
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Simple math evaluator expression
class SimpleEval:
    def __init__(self):
        self.operators = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.Pow: operator.pow,
            ast.USub: operator.neg,
        }
        
        self.functions = {
            'max': max,
            'min': min,
            'abs': abs,
            'sqrt': math.sqrt,
        }
    
    def eval(self, expr, variables=None):
        if variables is None:
            variables = {}
        
        node = ast.parse(expr, mode='eval')
        return self._eval(node.body, variables)
    
    def _eval(self, node, variables):
        if isinstance(node, ast.Num):  # Number
            return node.n
        elif isinstance(node, ast.Name):  # Variable
            return variables[node.id]
        elif isinstance(node, ast.BinOp):  # Binary operation
            return self.operators[type(node.op)](
                self._eval(node.left, variables),
                self._eval(node.right, variables)
            )
        elif isinstance(node, ast.UnaryOp):  # Unary operation
            return self.operators[type(node.op)](
                self._eval(node.operand, variables)
            )
        elif isinstance(node, ast.Call):  # Function call
            func = self.functions[node.func.id]
            args = [self._eval(arg, variables) for arg in node.args]
            return func(*args)
        else:
            raise TypeError(f"Unsupported operation: {type(node)}")

# Extract tecplot variables using regular expression patten
def extract_variables(variable_string):
    pattern = r'"(.*?)"(?=\s*,|\s*$)'
    return re.findall(pattern, variable_string)

# extract specific data from Tecplot Ascii format
def extract_tecplot_data(file_path, target_time, target_variable):
    """
    Extracts specific data from Tecplot ASCII format.
    Finds exact time match or closest available time.
    
    Args:
        file_path: Path to the data file
        target_time: Time value to match (as float or scientific notation string)
        target_variable: Exact variable name to extract
    
    Returns:
        A tuple containing (closest_time, value) if found, (None, None) otherwise
    """
    # Read the file
    with open(file_path, 'r') as f:
        lines = f.readlines()
    
    # Parse variables
    variables_line = None
    for line in lines:
        if line.startswith('variables ='):
            variables_line = line
            break
    
    if not variables_line:
        raise ValueError("Variables line not found in file")
    
    # Clean and extract variable names
    variables = extract_variables(variables_line)
    
    try:
        var_index = variables.index(target_variable)
    except ValueError:
        raise ValueError(f"Variable '{target_variable}' not found in dataset '{file_path}'")
    
    # Find data zone and search for best time match
    in_data_zone = False
    best_match = None
    min_time_diff = float('inf')
    target_time_float = float(target_time)
    
    for line in lines:
        # Skip comments and metadata
        if line.startswith('!') or line.startswith('title') or line.startswith('variables'):
            continue
        
        # Detect start of data zone
        if line.startswith('zone'):
            in_data_zone = True
            continue
        
        if in_data_zone and line.strip():
            # Split data line into values
            #values = line.split()
            values = re.split(r'[,\s]+', line.strip())
            if not values:
                continue
                
            try:
                current_time = float(values[0])
                current_value = float(values[var_index])
                
                # Calculate time difference
                time_diff = abs(current_time - target_time_float)
                
                # Check for exact match
                if time_diff == 0:
                    #print('var_name: '+str(target_variable)+', var_index: '+str(var_index)+', value: '+str(current_value))
                    return (current_time, current_value)
                
                # Track closest time
                if time_diff < min_time_diff:
                    min_time_diff = time_diff
                    best_match = (current_time, current_value)
                    
            except (IndexError, ValueError):
                continue
    
    return best_match if best_match else (None, None)

# Extract string value that is enclosed by ''
def get_keyword(line):
    extracted = ''
    try:
        start = line.index("'") + 1   # Find first quote and move to next char
        end = line.index("'", start)  # Find closing quote after start
        extracted = line[start:end]
    except ValueError:
        extracted = ''
    return extracted

# Extract the first number from a given line before the delimiter
def get_keyword_before(line,char):
    extracted = line
    try:
        idx = line.index(char)  # Find closing quote after start
        if (idx >= 0):
            extracted = line[0:idx]
    except ValueError:
        extracted = ''
    return extracted

# Extract the first number from a given line after the delimiter
def get_keyword_after(line,char):
    extracted = line
    try:
        idx = line.index(char)  # Find closing quote after start
        if (idx > 0):
            extracted = line[idx+1:]
    except ValueError:
        extracted = ''
    return extracted

# Read CDR template file
def read_cdr-template_file(file_path):
    """
    Reads a text file with multiple groups following this format for each group:
    - First line: keyword enclosed in '', e.g., 'keyword: a'
    - Second line: number of output variables
    - Following lines: data rows with four columns enclosed in ''
    - Groups are separated by 'keyword: a', e.g., 'cdr result: co2'
    - Lines starting with ! are comments and should be skipped
    
    Returns:
    - List of tuples, each containing:
      (keyword (str), num_variables (int), data (2D list of strings))
    """
    groups = []
    current_group = []
    
    with open(file_path, 'r') as file:
        for line in file:
            stripped_line = line.strip()
          
            # Skip comments and blank lines
            if not stripped_line or stripped_line.startswith('!'):
                continue
            elif stripped_line.startswith("'end group'"):
                if current_group:  # Blank line after content signals group end
                    groups.append(current_group)
                    current_group = []
                continue
            else:
                current_group.append(stripped_line)
        
        # Add the last group if file doesn't end with blank line
        if current_group:
            groups.append(current_group)
    
    # Process each group
    result = []
    for group in groups:
        if len(group) < 2:  # Need at least keyword and num_variables
            continue
            
        keyword = get_keyword(group[0])
        
        try:
            num_variables = int(get_keyword_before(group[1]," "))
        except ValueError:
            num_variables = 0
        
        data = []
        for line in group[2:]:
            columns = [col.strip("'") for col in line.split("'") if col.strip()]
            iend = min(len(columns),4)  # Only the first 4 columns are required
            if iend >= 2:
                data.append(columns[:iend])
        
        result.append((keyword, num_variables, data))
    
    return result


def main():

    ################################################
    # User specified input, change this accordingly.
    ################################################
    # file_path_cdr-template = 'cdr-template.cdr'
    # file_path_cdr_output = 'cdr-result.xlsx'
    # prefix_min3p_input = 'cdr-simulation_'
    # target_times = [1.0, 2.0, 3.0, 4.0]
    # time_unit = 'time [years]'
    
    # Example command: 
    # python cdr-template.py cdr-template.cdr cdr-result.xlsx cdr-simulation "1.0, 2.0, 3.0, 4.0" "time [years]"
    # python cdr-template.py cdr-template.cdr ./subdomain_2/cdr-result.xlsx ./subdomain_2/cdr-simulation "1.0, 2.0, 3.0, 4.0" "time [years]"
    
    if len(sys.argv) > 5:
    	file_path_cdr-template = sys.argv[1]
    	file_path_cdr_output = sys.argv[2]
    	prefix_min3p_input = sys.argv[3]+"_"
    	# Split on either spaces or commas (with optional surrounding whitespace)
    	target_times = [float(x) for x in re.split(r'[,\s]+', sys.argv[4].strip())]
    	time_unit = sys.argv[5]
    else:
    	print("Error in python arguments, please check")


    # Math expression evaluation
    evaluator = SimpleEval()

    # Initialize variables for excel output
    wb = Workbook()
    ws = wb.active

    # Set column width to 20 characters for the first 100 columns
    for col in range(1, 60):  # Adjust range as needed
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Create alignment style with text wrapping
    wrap_alignment = Alignment(wrap_text=True)
    current_row = 1
       
    # Data processing
    result_groups = read_cdr-template_file(file_path_cdr-template)
    for i, (keyword, num_variables, variables) in enumerate(result_groups, 1):
        print('\nOutput of '+str(keyword))
        # Create a dictionary to store variables
        var_name = ['unknown' for j in range(num_variables)]
        ivar = 0
        for row in variables:            
            if (row[0].find('=') > 0):
                var_name[ivar] = get_keyword_before(row[0],'=')
            else:
                var_name[ivar] = get_keyword_before(row[0],':')
            ivar = ivar + 1
        var_value = {name: None for name in var_name}  # Initialize with None
        
        # Variable to store result for each group
        data_head = np.array([f"item_{i}" for i in range(num_variables+1)], dtype='U128')  # Unicode string array
        data_head[0] = time_unit
        ivar = 0
        for row in variables:   
            ivar = ivar + 1         
            data_head[ivar] = str(row[1])

        data_value = np.zeros((len(target_times),num_variables+1))    

        # Write excel group header
        ws.cell(row=current_row, column=1, value=str(keyword))
        ws.cell(row=current_row, column=1).alignment = wrap_alignment
        current_row += 1
        
        # Write excel variables names
        for icol, str_var in enumerate(data_head, 1):
            ws.cell(row=current_row, column=icol, value=str_var)
            ws.cell(row=current_row, column=icol).alignment = wrap_alignment
        current_row += 1
        
        # Loop over target times
        for itime in range(len(target_times)):
            print('\nOutput target time: '+str(target_times[itime]))

            # Loop over all the variables to extract all directly available data
            ivar = 0
            for row in variables:            
                if (len(row) >= 4):
                    file_path_data = prefix_min3p_input+str(row[2])
                    target_time = target_times[itime]
                    target_variable = str(row[3])
                    closest_time, value = extract_tecplot_data(file_path_data, target_time, target_variable)
                    if closest_time is not None:                    
                        var_value[var_name[ivar]] = value
                        print('Direct output of '+str(row[1])+': '+str(value))
                    else:
                        print('No matching data found')
                ivar = ivar + 1 

            # Loop over all the variables to calculate all derived available data
            ivar = 0
            for row in variables:            
                expression = get_keyword_after(str(row[0]),'=')
                value = evaluator.eval(str(expression), var_value)
                var_value[var_name[ivar]] = value
                print('Derived output of '+str(row)+': '+str(value))
                ivar = ivar + 1

            # Reconstruct the output dataset
            data_value[itime][0] = target_times[itime]
            for ivar in range(num_variables):
                data_value[itime][ivar+1] = var_value[var_name[ivar]]

            # Write excel data
            for icol, value in enumerate(data_value[itime], 1):
                ws.cell(row=current_row, column=icol, value=value)
            current_row += 1

        # Add spacing between groups
        current_row += 2

    # Save result to file
    wb.save(file_path_cdr_output)

if __name__ == "__main__":
    main()